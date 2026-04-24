#!/usr/bin/env python3
"""
Conversion des données de sélection française → CSV compatible dashboard
Source : fichier Excel FFCK (format issu du logiciel de chronométrage)
Produit : data/french/fra_selection_AAAA-MM-JJ.csv

Usage : python scripts/prepare_fra_data.py <fichier.xlsx> [--output data/french/]
"""

import re
import csv
import sys
import argparse
from pathlib import Path
from datetime import date
from openpyxl import load_workbook

# ─── MAPPING DES COURSES ──────────────────────────────────────────────────────
# Extrait boat_class, gender, distance depuis le nom de course
COURSE_PATTERN = re.compile(
    r'FINALE\s+\w+\s+(K|C)(\d)\s*(HO|DO|MIX)\s+(\d[\d.,]*)\s*m',
    re.IGNORECASE
)

GENDER_MAP = {"HO": "H", "DO": "F", "MIX": "MIX"}

def parse_course(course_str):
    """Extrait (boat_class, gender, distance_m) du nom de course."""
    m = COURSE_PATTERN.search(str(course_str))
    if not m:
        return None, None, None
    boat_type  = m.group(1).upper()
    crew_size  = m.group(2)
    gender     = GENDER_MAP.get(m.group(3).upper(), m.group(3))
    dist_raw   = m.group(4).replace(',', '').replace('.', '')
    boat_class = f"{boat_type}{crew_size}"
    try:
        distance = int(dist_raw)
    except ValueError:
        distance = None
    return boat_class, gender, distance

def parse_temps(t):
    """Convertit '01:51,2' ou '1:51.20' en secondes flottantes."""
    if t is None:
        return None
    t = str(t).strip().replace(',', '.')
    m = re.match(r'(\d+):(\d+(?:\.\d+)?)', t)
    if m:
        return round(int(m.group(1)) * 60 + float(m.group(2)), 3)
    try:
        return float(t)
    except ValueError:
        return None

def fmt_temps(s):
    if s is None: return ""
    m = int(s) // 60
    sec = s - m * 60
    return f"{m}:{sec:05.2f}"

# ─── LECTURE EXCEL ────────────────────────────────────────────────────────────
def lire_excel(filepath):
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    rows_out = []
    # Détecter l'ordre des colonnes depuis le header
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip().lower() if c else "" for c in row]
            continue

        if not row[0]:
            continue

        r = dict(zip(headers, row))

        course     = r.get("course", "")
        place      = r.get("place", None)
        nom        = r.get("nom", "")
        annee      = r.get("année", r.get("annee", ""))
        categorie  = r.get("catégorie", r.get("categorie", ""))
        club       = r.get("club", "")
        temps_raw  = r.get("temps", None)

        boat_class, gender, distance = parse_course(course)
        if boat_class is None:
            continue

        time_s = parse_temps(temps_raw)
        if time_s is None:
            continue

        try:
            rank = int(place)
        except (TypeError, ValueError):
            continue

        rows_out.append({
            "course":      course,
            "boat_class":  boat_class,
            "gender":      gender,
            "distance_m":  distance,
            "rank":        rank,
            "athlete":     str(nom).strip(),
            "annee_nais":  annee,
            "categorie":   categorie,
            "club":        str(club).strip(),
            "time_raw":    fmt_temps(time_s),
            "time_seconds": time_s,
        })

    return rows_out

# ─── CALCUL DES MÉTRIQUES FRANÇAISES ─────────────────────────────────────────
def enrichir(rows):
    """
    Pour chaque ligne, calcule :
    - pct_gap_winner_fra : % d'écart au vainqueur français de la même épreuve
    """
    # Indexer les vainqueurs français par (boat_class, gender, distance)
    winners = {}
    for r in rows:
        if r["rank"] == 1:
            key = (r["boat_class"], r["gender"], r["distance_m"])
            winners[key] = r["time_seconds"]

    enriched = []
    for r in rows:
        r = r.copy()
        key = (r["boat_class"], r["gender"], r["distance_m"])
        w = winners.get(key)
        if w and w > 0:
            r["pct_gap_winner_fra"] = round(
                (r["time_seconds"] - w) / w * 100, 3
            )
        else:
            r["pct_gap_winner_fra"] = None
        enriched.append(r)
    return enriched

# ─── ÉCRITURE CSV ─────────────────────────────────────────────────────────────
FIELDNAMES = [
    "course", "boat_class", "gender", "distance_m",
    "rank", "athlete", "annee_nais", "categorie", "club",
    "time_raw", "time_seconds", "pct_gap_winner_fra",
]

def ecrire_csv(rows, output_path):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

# ─── RÉSUMÉ CONSOLE ───────────────────────────────────────────────────────────
def afficher_resume(rows):
    from collections import defaultdict
    par_epreuve = defaultdict(list)
    for r in rows:
        key = f"{r['boat_class']} {r['gender']} {r['distance_m']}m"
        par_epreuve[key].append(r)

    print(f"\n{'═'*60}")
    print(f"  Résumé des données françaises importées")
    print(f"{'═'*60}")
    print(f"  Total athlètes : {len(rows)}")
    print(f"  Épreuves       : {len(par_epreuve)}")
    print()
    for epreuve, athletes in sorted(par_epreuve.items()):
        winner = next((a for a in athletes if a["rank"] == 1), None)
        print(f"  {epreuve:<20} → {len(athletes)} athlètes | "
              f"Vainqueur : {winner['athlete'] if winner else '?'} "
              f"({winner['time_raw'] if winner else '?'})")
    print()

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Convertit l'Excel de sélection française en CSV dashboard")
    parser.add_argument("fichier", help="Chemin vers le fichier Excel (.xlsx)")
    parser.add_argument("--output", default="data/french", help="Dossier de sortie")
    args = parser.parse_args()

    filepath = Path(args.fichier)
    if not filepath.exists():
        print(f"❌ Fichier introuvable : {filepath}")
        sys.exit(1)

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f"fra_selection_{date.today()}.csv"

    print(f"Lecture de {filepath}…")
    rows = lire_excel(filepath)
    print(f"  {len(rows)} entrées lues")

    rows = enrichir(rows)
    afficher_resume(rows)

    ecrire_csv(rows, output_file)
    print(f"✅ CSV généré : {output_file}")
    print(f"\nPour utiliser dans le dashboard :")
    print(f"  → Section 4 du dashboard → Charger {output_file.name}")

if __name__ == "__main__":
    main()
